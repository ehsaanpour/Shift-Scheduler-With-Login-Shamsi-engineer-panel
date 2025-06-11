from flask import Flask, request, jsonify, render_template, redirect, url_for, session, flash, send_file
import json
import os
import calendar
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import shutil
import tempfile
import hashlib
import secrets
import jdatetime as jdt
import locale
import logging # Add logging import
from functools import wraps # Add this import at the top
import uuid # For unique notification IDs

app = Flask(__name__, template_folder='templates', static_folder='static')
# IMPORTANT: Set a fixed secret key. 
# For production, use an environment variable: app.secret_key = os.environ.get('SECRET_KEY')
# For development, you can use a hardcoded string (change this!):
app.secret_key = '940e26b1bfcbe2c0111ce0aaf3230d43' # Set a persistent secret key
# app.secret_key = secrets.token_hex(16)  # Commented out random generation

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Set jdatetime locale to Persian
try:
    # Ensure the locale exists on the system; this might vary
    # Use 'fa_IR.UTF-8' or similar if 'fa_IR' doesn't work
    locale.setlocale(locale.LC_ALL, 'fa_IR') 
except locale.Error:
    print("Warning: Persian locale 'fa_IR' not found. Using default locale.")
    # Fallback or handle error as needed
jdt.set_locale(jdt.FA_LOCALE)
PERSIAN_MONTH_NAMES = [
    "", "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]
PERSIAN_DAY_NAMES = [
    "شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنجشنبه", "جمعه"
]

# Global state for engineer panel lock
ENGINEER_PANEL_LOCKED = False
LOCK_MESSAGE = "پنل کاربری برای اعمال محدودیت‌های شما به دلیل عدم رعایت زمان‌بندی بسته شده است. لطفاً در زمان مشخص‌شده وارد شوید، در غیر این صورت با مدیر واحد تماس بگیرید"

# Configuration
DATA_DIR = '/data'  # Use absolute path as specified in liara.json
logger.info(f"Using DATA_DIR: {DATA_DIR}")
if not os.path.exists(DATA_DIR):
    logger.info(f"Creating DATA_DIR: {DATA_DIR}")
    os.makedirs(DATA_DIR)
    logger.info(f"Created DATA_DIR with permissions: {oct(os.stat(DATA_DIR).st_mode)[-3:]}")

ENGINEERS_FILE = os.path.join(DATA_DIR, 'engineers.json')
SCHEDULES_FILE = os.path.join(DATA_DIR, 'schedules.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
NOTIFICATIONS_FILE = os.path.join(DATA_DIR, 'notifications.json')
MESSAGES_FILE = os.path.join(DATA_DIR, 'messages.json')

# Log file permissions and existence
for file_path in [ENGINEERS_FILE, SCHEDULES_FILE, USERS_FILE, NOTIFICATIONS_FILE, MESSAGES_FILE]:
    if os.path.exists(file_path):
        logger.info(f"File {file_path} exists with permissions: {oct(os.stat(file_path).st_mode)[-3:]}")
    else:
        logger.info(f"File {file_path} does not exist")

WORKPLACES = ["Studio Hispan", "Studio Press", "Nodal", "Engineer Room"]
SHIFTS = ["Shift 1", "Shift 2", "Shift 3"]

# Initialize data files if they don't exist
if not os.path.exists(ENGINEERS_FILE):
    with open(ENGINEERS_FILE, 'w') as f:
        json.dump([], f)

if not os.path.exists(SCHEDULES_FILE):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump({}, f)

if not os.path.exists(USERS_FILE):
    with open(USERS_FILE, 'w') as f:
        # Create a default admin user with new structure
        default_admin = {
            "username": "admin",
            "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
            "role": "admin",      # Use 'role' instead of 'is_admin'
            "engineer_name": None # Admins are not linked to an engineer
        }
        json.dump([default_admin], f)

# --- Initialize Notifications File ---
if not os.path.exists(NOTIFICATIONS_FILE):
    with open(NOTIFICATIONS_FILE, 'w') as f:
        json.dump([], f) # Initialize with an empty list

# --- Initialize Messages File ---
if not os.path.exists(MESSAGES_FILE):
    with open(MESSAGES_FILE, 'w') as f:
        json.dump([], f) # Initialize with an empty list

# Authentication functions
def load_users():
    try:
        with open(USERS_FILE, 'r') as f:
            users = json.load(f)
            # --- Migration: Ensure old users have new fields ---
            updated_users = []
            made_changes = False
            for user in users:
                if 'role' not in user:
                    user['role'] = 'admin' if user.get('is_admin', False) else 'engineer'
                    user['engineer_name'] = user.get('engineer_name', None) # Keep if exists, else None
                    if 'is_admin' in user: del user['is_admin'] # Remove old field
                    made_changes = True
                updated_users.append(user)
            if made_changes:
                save_users(updated_users) # Save migrated data
                return updated_users
            # --- End Migration ---
            return users
    except Exception as e:
        print(f"Error loading or migrating users: {e}")
        # Attempt to create default if file was invalid/empty
        if not os.path.exists(USERS_FILE) or os.path.getsize(USERS_FILE) == 0:
             with open(USERS_FILE, 'w') as f:
                default_admin = {
                    "username": "admin",
                    "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
                    "role": "admin",
                    "engineer_name": None
                }
                json.dump([default_admin], f)
                return [default_admin]
        return []

def save_users(users):
    # Ensure all users have the correct fields before saving
    validated_users = []
    for user in users:
        validated_user = {
            "username": user.get("username"),
            "password_hash": user.get("password_hash"),
            "role": user.get("role", "engineer"), # Default to engineer if missing
            "engineer_name": user.get("engineer_name", None)
        }
        # Ensure required fields are present
        if not validated_user["username"] or not validated_user["password_hash"]:
            print(f"Warning: Skipping invalid user data during save: {user}")
            continue
        validated_users.append(validated_user)
        
    with open(USERS_FILE, 'w') as f:
        json.dump(validated_users, f, indent=2) # Add indent for readability

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, password_hash):
    return hash_password(password) == password_hash

def get_user(username):
    users = load_users()
    for user in users:
        if user["username"] == username:
            return user
    return None

def authenticate_user(username, password):
    user = get_user(username)
    if not user:
        return None # Return None instead of False
    if not verify_password(password, user["password_hash"]):
        return None # Return None instead of False
    # Return the full user object (which now includes role and engineer_name)
    return user 

# Helper functions
def load_engineers():
    try:
        with open(ENGINEERS_FILE, 'r', encoding='utf-8') as f:
            engineers = json.load(f)
            print(f"LOAD_ENGINEERS: Successfully loaded {len(engineers)} engineers from file")
            
            # --- Data Cleaning/Validation --- 
            cleaned_engineers = []
            data_was_cleaned = False
            for eng in engineers:
                limits = eng.get('limitations')
                is_valid = True
                eng_name = eng.get('name', '[Unknown Engineer]')

                if limits is None:
                    eng['limitations'] = {}
                    # No cleaning needed, just ensuring it exists
                elif not isinstance(limits, dict):
                    print(f"LOAD_ENGINEERS CLEANUP: Limitations for {eng_name} was not a dict ({type(limits)}). Resetting to {{}}.")
                    is_valid = False
                else: # It is a dictionary, validate structure
                    for year_key, year_val in limits.items():
                        # Check Year Key: Must be a 4-digit string
                        if not isinstance(year_key, str) or not year_key.isdigit() or len(year_key) != 4:
                            print(f"LOAD_ENGINEERS CLEANUP: Invalid year key '{year_key}' ({type(year_key)}) found for {eng_name}. Resetting limitations.")
                            is_valid = False; break
                        
                        # Check Year Value: Must be a dictionary
                        if not isinstance(year_val, dict):
                            print(f"LOAD_ENGINEERS CLEANUP: Value for year '{year_key}' is not a dict ({type(year_val)}) for {eng_name}. Resetting limitations.")
                            is_valid = False; break
                            
                        # Check Month Keys/Values within the year
                        for month_key, month_val in year_val.items():
                            # Check Month Key: Must be a 1 or 2 digit string representing 1-12
                            is_valid_month_key = (isinstance(month_key, str) and 
                                                  month_key.isdigit() and 
                                                  1 <= int(month_key) <= 12)
                            if not is_valid_month_key:
                                print(f"LOAD_ENGINEERS CLEANUP: Invalid month key '{month_key}' ({type(month_key)}) in year '{year_key}' for {eng_name}. Resetting limitations.")
                                is_valid = False; break
                                
                            # Check Month Value: Must be a dictionary (containing day:shifts)
                            if not isinstance(month_val, dict):
                                print(f"LOAD_ENGINEERS CLEANUP: Value for month '{month_key}' in year '{year_key}' is not a dict ({type(month_val)}) for {eng_name}. Resetting limitations.")
                                is_valid = False; break
                                
                        if not is_valid: break # Exit year loop if month was invalid

                # If validation failed, reset the limitations for this engineer
                if not is_valid:
                    eng['limitations'] = {}
                    data_was_cleaned = True # Mark that we made a change
                
                cleaned_engineers.append(eng)
            # --- End Data Cleaning --- 
            
            # If we cleaned any data, save it back immediately
            if data_was_cleaned:
                print("LOAD_ENGINEERS CLEANUP: Invalid limitation structures found and reset. Saving cleaned data...")
                save_engineers(cleaned_engineers) # Save the cleaned list
                return cleaned_engineers # Return the cleaned data for immediate use
            else:
                return engineers # Return original data if no cleaning was needed

    except Exception as e:
        print(f"LOAD_ENGINEERS ERROR: {str(e)}")
        return []

def save_engineers(engineers):
    lock_file_path = ENGINEERS_FILE + '.lock'
    try:
        logger.info(f"Attempting to save engineers to {ENGINEERS_FILE}")
        logger.info(f"Current user: {os.getuid()}")
        logger.info(f"Current directory: {os.getcwd()}")
        
        # Basic file locking mechanism
        with open(lock_file_path, 'w') as lock_file:
            logger.info(f"Acquired lock {lock_file_path}")
            
            if not isinstance(engineers, list):
                logger.error(f"Engineers is not a list, it's a {type(engineers)}")
                return

            if len(engineers) == 0:
                logger.warning("Saving an empty engineers list!")
                
            # Create backup of current file if it exists
            backup_file = f"{ENGINEERS_FILE}.bak"
            if os.path.exists(ENGINEERS_FILE):
                logger.info(f"Creating backup at {backup_file}")
                shutil.copy2(ENGINEERS_FILE, backup_file)
                
            logger.info(f"Preparing to save {len(engineers)} engineers to file")
            
            # Write to a temporary file first
            temp_file_path = ENGINEERS_FILE + '.tmp'
            try:
                with open(temp_file_path, 'w', encoding='utf-8') as f:
                    json.dump(engineers, f, indent=2, ensure_ascii=False)
                    f.flush()
                    os.fsync(f.fileno())
                    logger.info(f"Data flushed and synced to temporary file {temp_file_path}")
                
                # Atomically replace the original file with the temporary file
                os.replace(temp_file_path, ENGINEERS_FILE)
                logger.info(f"Successfully saved {len(engineers)} engineers by replacing {ENGINEERS_FILE} with {temp_file_path}")
            except Exception as e:
                logger.error(f"Error during file operations: {str(e)}")
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                raise
    except Exception as e:
        logger.error(f"Error in save_engineers: {str(e)}")
        raise
    finally:
        if os.path.exists(lock_file_path):
            os.remove(lock_file_path)
            logger.info(f"Released lock {lock_file_path}")

def load_schedules():
    try:
        with open(SCHEDULES_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_schedules(schedules):
    with open(SCHEDULES_FILE, 'w') as f:
        json.dump(schedules, f)

# Login required decorator
def login_required(f):
    @wraps(f) # Use wraps
    def decorated_function(*args, **kwargs):
        logging.info(f"LOGIN_REQUIRED decorator for {f.__name__}: Checking session. Session data: {session}") # Added log
        if 'user' not in session:
            flash('لطفا برای دسترسی به این صفحه وارد شوید.', 'warning') # Translated
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    # decorated_function.__name__ = f.__name__ # No longer needed
    return decorated_function

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('لطفا برای دسترسی به این صفحه وارد شوید.', 'warning')
            return redirect(url_for('login_page'))
            
        # Get user from database to verify role
        user = get_user(session['user']['username'])
        if not user or user.get('role') != 'admin':
            flash('برای دسترسی به این صفحه به سطح دسترسی ادمین نیاز دارید.', 'danger')
            session.clear()  # Clear potentially compromised session
            return redirect(url_for('login_page'))
            
        return f(*args, **kwargs)
    return decorated_function

# Engineer required decorator
def engineer_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('لطفا برای دسترسی به این صفحه وارد شوید.', 'warning')
            return redirect(url_for('login_page'))
            
        # Get user from database to verify role
        user = get_user(session['user']['username'])
        if not user or user.get('role') != 'engineer':
            flash('برای دسترسی به این صفحه به سطح دسترسی مهندس نیاز دارید.', 'danger')
            session.clear()  # Clear potentially compromised session
            return redirect(url_for('login_page'))
            
        # Verify engineer is linked
        if not user.get('engineer_name'):
            flash('حساب کاربری شما به پروفایل مهندس متصل نیست. لطفا با ادمین تماس بگیرید.', 'danger')
            session.clear()
            return redirect(url_for('login_page'))
            
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
@login_required
def index():
    # --- Fetch and Mark Notifications for Admin --- 
    unread_notifications = []
    if session.get('user') and session['user'].get('role') == 'admin':
        admin_username = session['user']['username']
        unread_notifications = get_unread_notifications(admin_username)
        if unread_notifications: # Only mark as read if there were unread ones
            mark_notifications_as_read(admin_username)
    # --- End Notification Handling ---
    
    return render_template("index.html", 
                           workplaces=WORKPLACES, 
                           shifts=SHIFTS, 
                           username=session['user']['username'],
                           unread_notifications=unread_notifications # Pass notifications to template
                           )

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = authenticate_user(username, password)
        
        if not user:
            return render_template("login.html", error="Invalid username or password")
        
        # Clear any existing session data
        session.clear()
        
        # Store minimal user info in session
        session['user'] = {
            "username": user["username"],
            "role": user["role"],
            "engineer_name": user.get("engineer_name"),
            "login_time": datetime.now(timezone.utc).isoformat()
        }
        
        # Set session security options
        # session.permanent = False  # Session expires when browser closes <-- Commenting this out
        session.modified = True    # Force session update
        
        logging.info(f"LOGIN_SUCCESS: Session set for user {user['username']}. Role: {user['role']}")
        
        # Redirect based on role
        if user["role"] == "admin":
            flash(f"ورود ادمین موفقیت‌آمیز بود. خوش آمدید {user['username']}!", 'success')
            return redirect(url_for('index'))
        elif user["role"] == "engineer":
            if user.get('engineer_name'):
                flash(f"ورود مهندس موفقیت‌آمیز بود. خوش آمدید {user['engineer_name']}!", 'success')
                return redirect(url_for('engineer_dashboard'))
            else:
                session.clear()
                flash('ورود ناموفق: حساب کاربری مهندس به طور کامل پیکربندی نشده است. با ادمین تماس بگیرید.', 'danger')
                return redirect(url_for('login_page'))
        else:
            session.clear()
            flash('ورود ناموفق: نقش کاربر ناشناخته است.', 'danger')
            return redirect(url_for('login_page'))

    # If GET request
    if 'user' in session:
        # Verify session data against database
        user = get_user(session['user'].get('username'))
        if not user or user.get('role') != session['user'].get('role'):
            session.clear()
            flash('نشست نامعتبر است. لطفا دوباره وارد شوید.', 'warning')
            return redirect(url_for('login_page'))
            
        # Redirect based on role
        if user["role"] == "admin":
            return redirect(url_for('index'))
        elif user["role"] == "engineer" and user.get('engineer_name'):
            return redirect(url_for('engineer_dashboard'))
        else:
            session.clear()
            flash('نشست نامعتبر است. لطفا دوباره وارد شوید.', 'warning')
            return redirect(url_for('login_page'))

    return render_template("login.html", error=None)

@app.route('/logout')
def logout():
    # Clear session
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/admin')
@admin_required # Now checks for role='admin'
def admin_page():
    # --- Mark notifications as read for this admin ---
    admin_username = session['user']['username']
    mark_notifications_as_read(admin_username)
    # --- End Mark Notifications ---

    users = load_users()
    engineers = load_engineers() # Load engineers for linking
    
    # Get unread message counts for all engineers
    unread_admin_counts = get_unread_admin_message_counts()

    # Prepare users data, adding engineer link info and unread counts
    processed_users = []
    for user in users:
        engineer_exists = False
        if user.get('role') == 'engineer' and user.get('engineer_name'):
            engineer_exists = any(e['name'] == user['engineer_name'] for e in engineers)
        user_data = user.copy() # Avoid modifying original dict
        user_data['engineer_exists'] = engineer_exists
        # Add unread count for this engineer if they are an engineer
        if user.get('role') == 'engineer' and user.get('engineer_name'):
             user_data['unread_message_count'] = unread_admin_counts.get(user['engineer_name'], 0)
        else:
             user_data['unread_message_count'] = 0 # Admins don't have messages tied this way
             
        processed_users.append(user_data)

    return render_template('admin.html', 
                            users=processed_users, 
                            engineers=engineers, # Pass engineers for dropdown
                            username=session['user']['username'])

@app.route('/admin/users', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role', 'engineer') # Get role from form, default to engineer
    engineer_name = request.form.get('engineer_name') if role == 'engineer' else None

    if not username or not password:
         flash('Username and password are required.', 'danger')
         return redirect(url_for('admin_page'))

    if role == 'engineer' and not engineer_name:
         flash('An engineer must be selected when creating an engineer account.', 'danger')
         return redirect(url_for('admin_page'))

    users = load_users()

    # Check if username already exists
    if any(u['username'] == username for u in users):
        flash(f"Username '{username}' already exists.", 'danger')
        return redirect(url_for('admin_page'))

    # Check if engineer is already linked (if applicable)
    if role == 'engineer' and any(u.get('engineer_name') == engineer_name for u in users if u['role'] == 'engineer'):
         flash(f"Engineer '{engineer_name}' is already linked to another user account.", 'danger')
         return redirect(url_for('admin_page'))

    # Create new user
    password_hash = hash_password(password)
    new_user = {
        "username": username,
        "password_hash": password_hash,
        "role": role,
        "engineer_name": engineer_name
    }
    users.append(new_user)
    save_users(users)

    flash(f"{role.capitalize()} user '{username}' created successfully.", 'success')
    return redirect(url_for('admin_page'))

@app.route('/admin/users/<username>', methods=['DELETE'])
@admin_required
def delete_user(username):
    # Prevent deleting yourself
    if username == session['user']['username']:
        return jsonify({"error": "Cannot delete yourself"}), 400

    users = load_users()
    
    # --- Prevent deleting the last admin ---
    admin_users = [u for u in users if u['role'] == 'admin']
    user_to_delete = next((u for u in users if u['username'] == username), None)
    
    if user_to_delete and user_to_delete['role'] == 'admin' and len(admin_users) <= 1:
        return jsonify({"error": "Cannot delete the last admin user."}), 400
    # --- End last admin check ---

    original_count = len(users)
    users = [u for u in users if u["username"] != username]

    if len(users) < original_count:
        save_users(users)
        return jsonify({"status": "success", "message": f"User '{username}' deleted."})
    else:
        return jsonify({"error": f"User '{username}' not found."}), 404

@app.route('/excel')
@login_required
def excel_version():
    """
    Alternative version of the main page with improved Excel generation capabilities.
    Use this URL if the main page is having issues.
    """
    return render_template("index.html", 
                          workplaces=WORKPLACES, 
                          shifts=SHIFTS, 
                          username=session['user']['username'])

@app.route('/excel-generator')
@login_required
def excel_generator_page():
    """
    Standalone page for generating Excel files from the schedule.
    Use this if the main Excel generation functionality isn't working.
    """
    # Get current Jalali date
    now_gregorian = datetime.now()
    now_jalali = jdt.datetime.fromgregorian(datetime=now_gregorian)
    current_jalali_year = now_jalali.year
    current_jalali_month = now_jalali.month
    
    return render_template("excel_generator.html", 
                          username=session['user']['username'],
                          current_jalali_year=current_jalali_year,
                          current_jalali_month=current_jalali_month,
                          persian_month_names=PERSIAN_MONTH_NAMES # Pass month names
                          )

# API Routes
@app.route('/api/engineers', methods=['GET'])
@login_required
def get_engineers():
    return jsonify(load_engineers())

@app.route('/api/engineers', methods=['POST'])
@admin_required
def add_engineer():
    data = request.json
    print(f"ADD_ENGINEER: Received data for engineer: {data['name']}")
    
    # Make a local copy of all engineers to avoid reference issues
    engineers = load_engineers()
    print(f"ADD_ENGINEER: Loaded {len(engineers)} existing engineers")
    
    # Check if updating or adding new
    engineer_exists = False
    engineer_index = -1
    
    # First, find if the engineer exists and get its index
    for i, eng in enumerate(engineers):
        if eng['name'] == data['name']:
            engineer_exists = True
            engineer_index = i
            break
    
    if engineer_exists:
        print(f"ADD_ENGINEER: Updating existing engineer at index {engineer_index}: {data['name']}")
        # Get the existing engineer's data
        existing_engineer_data = engineers[engineer_index]

        # Check if year and month are provided for a specific limitation update
        year = data.get('year')
        month = data.get('month')

        # Initialize the updated data with existing values
        updated_engineer = existing_engineer_data.copy() 

        # Update basic info always
        updated_engineer['workplaces'] = data['workplaces']
        updated_engineer['minShifts'] = data.get('minShifts', existing_engineer_data.get('minShifts', 10))
        updated_engineer['maxShifts'] = data.get('maxShifts', existing_engineer_data.get('maxShifts', 30))

        if year is not None and month is not None:
            print(f"ADD_ENGINEER: Updating limitations for specific period: Year {year}, Month {month}")
            limitations_for_month = data.get('limitations', {})
            
            # --- Ensure consistent STRING key usage --- 
            year_str = str(year)
            month_str = str(month)
            
            # Ensure the main limitations dictionary exists and is a dict
            if not isinstance(updated_engineer.get('limitations'), dict):
                updated_engineer['limitations'] = {}
            
            # Ensure the year key (as STRING) exists and is a dict
            if year_str not in updated_engineer['limitations'] or not isinstance(updated_engineer['limitations'][year_str], dict):
                updated_engineer['limitations'][year_str] = {}
                
            # Update the limitations for the specific month (using string keys)
            updated_engineer['limitations'][year_str][month_str] = limitations_for_month
            # --- End STRING key usage correction ---
            
            print(f"ADD_ENGINEER: Updated limitations structure: {updated_engineer['limitations']}")
            
            # Add notification if limitations were updated
            try:
                add_limitation_update_notification(updated_engineer['name'], year_str, month_str)
                print(f"ADD_ENGINEER: Added limitation update notification for {updated_engineer['name']}")
            except Exception as e:
                print(f"ADD_ENGINEER: Error adding notification: {e}")
        else:
             # If year/month are not provided, check if 'limitations' key exists in request data
            if 'limitations' in data:
                 print(f"ADD_ENGINEER: Updating/Replacing ENTIRE limitations structure based on request data.")
                 updated_engineer['limitations'] = data['limitations']

        # Replace the old engineer with the updated one in the main list
        engineers[engineer_index] = updated_engineer
    else:
        print(f"ADD_ENGINEER: Adding new engineer: {data['name']}")
        # Add the new engineer
        new_engineer = {
            'name': data['name'],
            'workplaces': data['workplaces'],
            'limitations': data.get('limitations', {}),
            'minShifts': data.get('minShifts', 10),
            'maxShifts': data.get('maxShifts', 30)
        }
        engineers.append(new_engineer)
    
    print(f"ADD_ENGINEER: Final list contains {len(engineers)} engineers")
    print(f"ADD_ENGINEER: Engineer names: {[eng.get('name', 'UNNAMED') for eng in engineers]}")
    
    # Make sure we're saving a copy to avoid any reference issues
    save_engineers(engineers[:])
    
    # Verify engineers were saved correctly
    verification = load_engineers()
    print(f"ADD_ENGINEER: Verification loaded {len(verification)} engineers")
    print(f"ADD_ENGINEER: Verified engineer names: {[eng.get('name', 'UNNAMED') for eng in verification]}")
    
    return jsonify({"status": "success"})

@app.route('/api/engineers/<n>', methods=['DELETE'])
@admin_required
def delete_engineer(n):
    engineers = load_engineers()
    engineers = [eng for eng in engineers if eng['name'] != n]
    save_engineers(engineers)
    return jsonify({"status": "success"})

@app.route('/api/schedule', methods=['GET'])
@login_required
def get_schedule():
    # Default to current Jalali year and month
    now_gregorian = datetime.now()
    now_jalali = jdt.datetime.fromgregorian(datetime=now_gregorian)
    
    year = request.args.get('year', default=now_jalali.year, type=int)
    month = request.args.get('month', default=now_jalali.month, type=int)
    
    schedules = load_schedules()
    # Use Jalali year/month for the key
    period_key = f"{year}-{month}" 
    
    if period_key in schedules:
        return jsonify(schedules[period_key])
    return jsonify({})

@app.route('/api/schedule', methods=['POST'])
@admin_required
def save_schedule():
    data = request.json
    # Assume year/month received from frontend are Jalali
    year = data.get('year')
    month = data.get('month')
    # Get the complete workplaces data from the request
    workplaces_data_from_request = data.get('workplaces', {})

    schedules = load_schedules()

    # Use Jalali year/month for the key
    period_key = f"{year}-{month}" 

    # *** FIX: Directly replace the data for the period ***
    # Instead of iterating and merging, just assign the received data.
    # This ensures that if the frontend sends a complete (potentially empty)
    # structure for the month, it fully overwrites whatever was there before.
    schedules[period_key] = workplaces_data_from_request

    save_schedules(schedules)
    return jsonify({"status": "success"})

@app.route('/api/generate_excel', methods=['POST'])
@login_required
def generate_excel():
    data = request.json
    # Assume year/month received from frontend are Jalali
    year = data.get('year')
    month = data.get('month')
    
    schedules = load_schedules()
    # Use Jalali year/month for the key
    period_key = f"{year}-{month}"
    
    if period_key not in schedules:
        return jsonify({"error": "No schedule data found for selected period"}), 404
    
    # Generate Excel files for each workplace
    excel_files = []
    for workplace in WORKPLACES:
        # Use Jalali year/month in filename
        filename = f"{workplace.replace(' ', '_')}_{year}_{month}.xlsx"
        file_path = os.path.join(DATA_DIR, filename)
        
        # Create Excel with formatting using Jalali calendar info
        create_excel_schedule(file_path, workplace, year, month, schedules[period_key].get(workplace, {}))
        excel_files.append(filename)
    
    return jsonify({
        "status": "success",
        "files": excel_files
    })

@app.route('/api/download/<filename>')
@login_required
def download_file(filename):
    file_path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, as_attachment=True)

@app.route('/api/pattern/upload', methods=['POST'])
@admin_required
def upload_pattern():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Invalid file format. Only Excel files (.xlsx, .xls) are supported."}), 400
    
    # --- Revised Temp File Handling ---
    tmp_path = None
    workbook = None 
    try:
        # Create a temporary file path
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd) # Close the file descriptor, we just need the path

        # Save the uploaded file directly to the path
        file.save(tmp_path)
        
        # Open the Excel file
        workbook = openpyxl.load_workbook(tmp_path, data_only=True)
        sheet = workbook.active
        
        # Parse the pattern
        pattern = {}
        max_rows = min(sheet.max_row, 31)
        for day in range(1, max_rows + 1):
            pattern[str(day)] = {}
            for shift in range(1, min(4, sheet.max_column + 1)):
                cell_value = sheet.cell(row=day, column=shift).value
                if cell_value:
                    pattern[str(day)][f"shift{shift}"] = str(cell_value).strip()
        
        # Explicitly close the workbook *before* returning
        workbook.close()
        workbook = None # Indicate workbook is closed

        return jsonify({
            "status": "success", 
            "pattern": pattern
        })
    except Exception as e:
        logging.error(f"Error processing pattern file '{file.filename}': {e}", exc_info=True)
        # Ensure workbook is closed if an error occurred during processing
        if workbook: 
            try:
                workbook.close()
            except Exception as close_err:
                logging.error(f"Error closing workbook during exception handling: {close_err}")
        return jsonify({"error": f"Error processing Excel file: {str(e)}"}), 500
    finally:
        # Clean up the temporary file if the path was created
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception as remove_err:
                 # Log if removal fails, but don't crash the request
                logging.error(f"Failed to remove temporary file {tmp_path}: {remove_err}")

def create_excel_schedule(file_path, workplace, year, month, schedule_data):
    try:
        # Ensure year and month are integers
        year_int = int(year)
        month_int = int(month)

        # --- Calculate num_days manually to bypass potential .daysinmonth bug ---
        if 1 <= month_int <= 6:
            num_days = 31
        elif 7 <= month_int <= 11:
            num_days = 30
        elif month_int == 12:
            if jdt.isleap(year_int):
                num_days = 30
            else:
                num_days = 29
        else:
            raise ValueError(f"Invalid month number: {month_int}")
        # --- End manual calculation ---
        
        # Get month name 
        jalali_month_name = PERSIAN_MONTH_NAMES[month_int]

    except ValueError as ve:
        print(f"Error: Invalid integer year/month or invalid Jalali date: {year}-{month}. Details: {ve}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = f"ValueError for {year}-{month}: {ve}"
        wb.save(file_path)
        return
    # Removed other specific exception handlers (TypeError, AttributeError, generic Exception) related to the initial j_date_start
    # The code will now rely on Flask's default error handling if something unexpected happens after the ValueError check.

    # --- If we reach here, num_days and jalali_month_name should be set correctly ---

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{workplace} Schedule"
    
    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    weekend_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color="FFFFFF")
    centered = Alignment(horizontal='center', vertical='center')
    
    # Create title using Jalali month name and year
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    # Use the jalali_month_name obtained safely from the try block
    title_cell.value = f"{workplace} - {jalali_month_name} {year_int}" 
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = centered
    
    # Create headers
    headers = ["Day", "Shift 1", "Shift 2", "Shift 3"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = centered
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Fill in days (using num_days obtained safely from the try block)
    for day in range(1, num_days + 1):
        row = day + 3
        
        # Get Jalali date and Gregorian date for weekday calculation
        try:
            # Use integer year/month
            j_date = jdt.date(year_int, month_int, day) 
            
            g_date = j_date.togregorian()
            
            # Get Persian day name 
            gregorian_weekday = g_date.weekday() # Monday is 0, Sunday is 6
            
            # Correct mapping: Gregorian Mon(0)->Shamsi Mon(2), Sat(5)->Shamsi Sat(0), Sun(6)->Shamsi Sun(1)
            persian_weekday_index = (gregorian_weekday + 2) % 7 
            persian_day_name = PERSIAN_DAY_NAMES[persian_weekday_index]
            
        except ValueError as loop_ve:
            print(f"Error in day loop (ValueError): Could not process date {year_int}-{month_int}-{day}. Error: {loop_ve}")
            persian_day_name = "خطا"
            g_date = None 
        except Exception as loop_e: # Catch any unexpected error in the loop
            print(f"Error in day loop (Exception): Failed on day {day}. Error: {loop_e}, Type: {type(loop_e)}")
            # Optionally re-raise or handle more gracefully 
            # For now, just print and continue with error values
            persian_day_name = "خطا"
            g_date = None 
        
        # Day column with Persian day name
        day_cell = ws.cell(row=row, column=1)
        day_cell.value = f"{day} - {persian_day_name}" 
        day_cell.border = border
        day_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Weekend formatting
        # is_weekend = (g_date.weekday() >= 5) if g_date else False # Original line: Check if Gregorian weekday is 5 (Sat) or 6 (Sun)
        # Shamsi weekend: Thursday (Gregorian weekday 3) and Friday (Gregorian weekday 4)
        is_weekend = (g_date.weekday() == 3 or g_date.weekday() == 4) if g_date else False 
        if is_weekend:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = weekend_fill
        
        # Fill in shifts
        day_str = str(day)
        if day_str in schedule_data:
            for shift_idx, shift in enumerate(SHIFTS, 1):
                shift_key = f"shift{shift_idx}"
                if shift_key in schedule_data[day_str]:
                    cell = ws.cell(row=row, column=shift_idx + 1)
                    cell.value = schedule_data[day_str][shift_key]
                    cell.border = border
                    cell.alignment = centered
                else:
                    ws.cell(row=row, column=shift_idx + 1).border = border
        else:
            for shift_idx in range(1, 4):
                ws.cell(row=row, column=shift_idx + 1).border = border
    
    # Set row height
    for row in range(1, num_days + 5):
        ws.row_dimensions[row].height = 25
    
    # Save workbook
    wb.save(file_path)

# --- New Function to Create Overall Excel Schedule ---
def create_overall_excel_schedule(file_path, year, month, schedules_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Overall Shifts {year}-{month}"

    # Define styles (can be reused or adapted from create_excel_schedule)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    weekend_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF")
    centered = Alignment(horizontal='center', vertical='center')
    title_font = Font(bold=True, size=14)

    current_column_offset = 1 # Starting column for the first table

    try:
        year_int = int(year)
        month_int = int(month)
        jalali_month_name = PERSIAN_MONTH_NAMES[month_int]

        if 1 <= month_int <= 6:
            num_days = 31
        elif 7 <= month_int <= 11:
            num_days = 30
        elif month_int == 12:
            num_days = 30 if jdt.isleap(year_int) else 29
        else:
            raise ValueError(f"Invalid month: {month_int}")

    except ValueError as ve:
        logging.error(f"Error creating overall Excel: Invalid year/month - {ve}")
        ws.cell(row=1, column=1).value = f"Error: Invalid year/month provided ({year}-{month})."
        wb.save(file_path)
        return

    for workplace_index, workplace in enumerate(WORKPLACES):
        schedule_data_for_workplace = schedules_data.get(workplace, {})
        
        # --- Workplace Title ---
        title_start_col = current_column_offset
        title_end_col = current_column_offset + 3 # Shifts + Day column
        ws.merge_cells(start_row=1, start_column=title_start_col, end_row=1, end_column=title_end_col)
        title_cell = ws.cell(row=1, column=title_start_col)
        title_cell.value = f"{workplace} - {jalali_month_name} {year_int}"
        title_cell.font = title_font
        title_cell.alignment = centered

        # --- Headers ---
        headers = ["Day", "Shift 1", "Shift 2", "Shift 3"]
        for col_idx, header_text in enumerate(headers):
            cell = ws.cell(row=3, column=current_column_offset + col_idx)
            cell.value = header_text
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = centered
            ws.column_dimensions[get_column_letter(current_column_offset + col_idx)].width = 18

        # --- Fill in Days and Shifts ---
        for day in range(1, num_days + 1):
            current_row = day + 3
            try:
                j_date = jdt.date(year_int, month_int, day)
                g_date = j_date.togregorian()
                persian_weekday_index = (g_date.weekday() + 2) % 7
                persian_day_name = PERSIAN_DAY_NAMES[persian_weekday_index]
            except ValueError:
                persian_day_name = "خطا"
                g_date = None

            day_cell = ws.cell(row=current_row, column=current_column_offset)
            day_cell.value = f"{day} - {persian_day_name}"
            day_cell.border = border
            day_cell.alignment = Alignment(horizontal='left', vertical='center')

            is_weekend = (g_date.weekday() == 3 or g_date.weekday() == 4) if g_date else False
            if is_weekend:
                for col_offset in range(4): # Day + 3 shifts
                    ws.cell(row=current_row, column=current_column_offset + col_offset).fill = weekend_fill
            
            day_str = str(day)
            if day_str in schedule_data_for_workplace:
                for shift_idx, shift_name in enumerate(SHIFTS):
                    shift_key = f"shift{shift_idx + 1}"
                    cell_to_fill = ws.cell(row=current_row, column=current_column_offset + 1 + shift_idx)
                    if shift_key in schedule_data_for_workplace[day_str]:
                        cell_to_fill.value = schedule_data_for_workplace[day_str][shift_key]
                    cell_to_fill.border = border
                    cell_to_fill.alignment = centered
            else:
                for shift_idx in range(len(SHIFTS)):
                    ws.cell(row=current_row, column=current_column_offset + 1 + shift_idx).border = border
        
        # Set row height for this table section (approx num_days + 4 header/title rows)
        for r in range(1, num_days + 5):
            ws.row_dimensions[r].height = 22

        # Update column offset for the next table (4 columns for data + 1 for spacing)
        current_column_offset += (len(headers) + 1)

    wb.save(file_path)

# --- END New Function ---

# --- Engineer Dashboard Route ---
@app.route('/engineer/dashboard')
@engineer_required # Protect this route
def engineer_dashboard():
    global ENGINEER_PANEL_LOCKED, LOCK_MESSAGE # Access global lock state
    # Get engineer's name from session
    engineer_name = session.get('user', {}).get('engineer_name')
    if not engineer_name:
        flash('خطا: نام مهندس مرتبط یافت نشد.', 'danger')
        return redirect(url_for('login_page'))

    engineers = load_engineers()
    engineer = next((eng for eng in engineers if eng['name'] == engineer_name), None)

    if not engineer:
        flash(f'خطا: اطلاعات مهندس برای {engineer_name} یافت نشد.', 'danger')
        return redirect(url_for('login_page'))

    # Get current Jalali date for default selection
    now_gregorian = datetime.now()
    now_jalali = jdt.datetime.fromgregorian(datetime=now_gregorian)
    current_jalali_year = now_jalali.year
    current_jalali_month = now_jalali.month

    # Get unread message count for this engineer
    unread_message_count = get_unread_engineer_message_count(engineer_name)

    return render_template(
        'engineer_dashboard.html', 
        engineer=engineer, 
        username=session['user']['username'], # Pass username
        current_jalali_year=current_jalali_year,
        current_jalali_month=current_jalali_month,
        persian_month_names=PERSIAN_MONTH_NAMES,
        unread_message_count=unread_message_count, # Pass unread count
        is_panel_locked=ENGINEER_PANEL_LOCKED, # Pass lock status
        lock_message=LOCK_MESSAGE if ENGINEER_PANEL_LOCKED else "" # Pass lock message
    )

# --- API Endpoint for Engineer Saving Limitations ---
@app.route('/api/engineer/limitations', methods=['POST'])
@engineer_required
def save_engineer_limitations():
    engineer_name = session['user']['engineer_name']
    data = request.json
    new_month_limitations = data.get('limitations', {}) # Limits for the month being saved
    year = data.get('year')
    month = data.get('month')

    if not year or not month:
         return jsonify({"error": "Year and month are required to save limitations."}), 400

    try:
        # Convert to string keys immediately for consistency
        year_str = str(int(year))
        month_str = str(int(month))
    except (ValueError, TypeError):
         return jsonify({"error": "Invalid year or month format provided."}), 400

    print(f"Saving limitations for engineer: {engineer_name}, Year: {year_str}, Month: {month_str}")

    engineers = load_engineers()
    engineer_found = False
    updated_engineer_limitations_full = None # For logging the full structure

    for eng in engineers:
        if eng['name'] == engineer_name:
            # --- CORRECTED LOGIC START ---
            # Get the existing limitations structure, ensure it's a dict
            current_eng_limits = eng.get('limitations', {}) 
            if not isinstance(current_eng_limits, dict):
                print(f"Warning: Limitations for {engineer_name} was not a dict, resetting.")
                current_eng_limits = {} # Initialize/reset if not a dict
                
            # Ensure the year key (as string) exists
            if year_str not in current_eng_limits:
                current_eng_limits[year_str] = {}
            elif not isinstance(current_eng_limits[year_str], dict):
                print(f"Warning: Limitations for {engineer_name} year {year_str} was not a dict, resetting year.")
                current_eng_limits[year_str] = {} # Reset year if not a dict
                
            # Directly replace the limitations for THIS specific month (using string keys)
            current_eng_limits[year_str][month_str] = new_month_limitations
            
            # Update the engineer's limitations field in the main list
            eng['limitations'] = current_eng_limits 
            updated_engineer_limitations_full = current_eng_limits # Store the updated full structure for logging
            # --- CORRECTED LOGIC END ---
            engineer_found = True
            break
    
    if not engineer_found:
        return jsonify({"error": "Engineer profile not found."}), 404

    print(f"LIMIT_SAVE (Engineer): Updated full limitations structure for {engineer_name}: {updated_engineer_limitations_full}")
    save_engineers(engineers)
    
    # Add Notification for Admin (using string year/month)
    try:
        add_limitation_update_notification(engineer_name, year_str, month_str)
        print(f"LIMIT_SAVE (Engineer): Added limitation update notification for {engineer_name}")
    except Exception as e:
        print(f"LIMIT_SAVE (Engineer): Error adding notification: {e}")
    
    return jsonify({"status": "success"})

# --- Notification Helper Functions ---
def load_notifications():
    try:
        with open(NOTIFICATIONS_FILE, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return [] # Return empty list on error

def save_notifications(notifications):
    with open(NOTIFICATIONS_FILE, 'w') as f:
        json.dump(notifications, f, indent=2)

def add_limitation_update_notification(engineer_name, year, month):
    notifications = load_notifications()
    month_name = PERSIAN_MONTH_NAMES[int(month)] if 1 <= int(month) <= 12 else f"Month {month}"
    message = f"Engineer '{engineer_name}' updated limitations for {month_name} {year}."
    
    new_notification = {
        "id": str(uuid.uuid4()), # Unique ID
        "timestamp": datetime.utcnow().isoformat(), # Corrected: Removed extra datetime.
        "type": "limitations_updated",
        "engineer_name": engineer_name,
        "year": year, # Store year
        "month": month, # Store month
        "message": message, # Store pre-formatted message
        "read_by": [] # List of admin usernames who have read it
    }
    notifications.append(new_notification)
    # Keep only the latest N notifications if desired (e.g., 50)
    max_notifications = 50
    if len(notifications) > max_notifications:
        notifications = notifications[-max_notifications:]
        
    save_notifications(notifications)

def get_unread_notifications(admin_username):
    notifications = load_notifications()
    unread = []
    for notification in reversed(notifications): # Show newest first
        if notification.get("type") == "limitations_updated" and \
           admin_username not in notification.get("read_by", []):
            unread.append(notification)
    return unread

def mark_notifications_as_read(admin_username):
    notifications = load_notifications()
    updated = False
    for notification in notifications:
        if (notification.get("type") == "limitations_updated" and 
            admin_username not in notification.get("read_by", [])):
             notification["read_by"].append(admin_username)
             updated = True
    if updated:
        save_notifications(notifications)

# --- Messaging Helper Functions ---
def load_messages():
    try:
        with open(MESSAGES_FILE, 'r', encoding='utf-8') as f:
            # Handle empty file case
            content = f.read()
            if not content:
                return []
            return json.loads(content)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error loading messages: {e}")
        return [] # Return empty list on error

def save_messages(messages):
    try:
        with open(MESSAGES_FILE, 'w', encoding='utf-8') as f:
            json.dump(messages, f, indent=2, ensure_ascii=False)
    except IOError as e:
        print(f"Error saving messages: {e}")

def add_message(engineer_name, sender_role, content):
    messages = load_messages()
    new_message = {
        "message_id": str(uuid.uuid4()),
        "engineer_name": engineer_name, # Always the engineer involved
        "sender_role": sender_role, # 'engineer' or 'admin'
        "timestamp": datetime.now(timezone.utc).isoformat(), 
        "content": content,
        "read_by_engineer": sender_role == 'engineer', # Read by sender initially
        "read_by_admin": sender_role == 'admin'    # Read by sender initially
    }
    messages.append(new_message)
    save_messages(messages)
    return new_message

def get_messages_for_engineer(engineer_name):
    messages = load_messages()
    engineer_messages = [m for m in messages if m.get('engineer_name') == engineer_name]
    # Sort by timestamp ascending
    engineer_messages.sort(key=lambda x: x.get('timestamp', ''))
    return engineer_messages

def mark_engineer_messages_as_read(engineer_name):
    messages = load_messages()
    updated = False
    for msg in messages:
        # Mark messages *sent by admin* to this engineer as read
        if msg.get('engineer_name') == engineer_name and \
           msg.get('sender_role') == 'admin' and \
           not msg.get('read_by_engineer', False):
            msg['read_by_engineer'] = True
            updated = True
    if updated:
        save_messages(messages)
    return updated

def mark_admin_messages_as_read(engineer_name):
    messages = load_messages()
    updated = False
    for msg in messages:
        # Mark messages *sent by this engineer* as read by admin
        if msg.get('engineer_name') == engineer_name and \
           msg.get('sender_role') == 'engineer' and \
           not msg.get('read_by_admin', False):
            msg['read_by_admin'] = True
            updated = True
    if updated:
        save_messages(messages)
    return updated

def get_unread_engineer_message_count(engineer_name):
    messages = load_messages()
    count = 0
    for msg in messages:
        if msg.get('engineer_name') == engineer_name and \
           msg.get('sender_role') == 'admin' and \
           not msg.get('read_by_engineer', False):
            count += 1
    return count

def get_unread_admin_message_counts(): # Ensure this function is defined
    messages = load_messages()
    counts = {}
    for msg in messages:
        # Count unread messages sent *by engineers*
        if msg.get('sender_role') == 'engineer' and \
           not msg.get('read_by_admin', False):
            eng_name = msg.get('engineer_name')
            if eng_name:
                counts[eng_name] = counts.get(eng_name, 0) + 1
    return counts

# --- END Messaging Helper Functions ---

# --- Messaging API Routes ---
@app.route('/api/messages/send', methods=['POST'])
@login_required
def send_message_api():
    data = request.get_json()
    if not data or 'content' not in data:
        return jsonify({"status": "error", "message": "محتوای پیام الزامی است"}), 400

    content = data['content'].strip()
    if not content:
        return jsonify({"status": "error", "message": "محتوای پیام نمی‌تواند خالی باشد"}), 400

    user = session['user']
    sender_role = user['role']
    engineer_name = None

    if sender_role == 'engineer':
        engineer_name = user.get('engineer_name')
        if not engineer_name:
             return jsonify({"status": "error", "message": "کاربر مهندس معتبر نیست"}), 403
    elif sender_role == 'admin':
        # Admin must specify which engineer they are messaging
        engineer_name = data.get('engineer_name')
        if not engineer_name:
            return jsonify({"status": "error", "message": "نام مهندس برای ارسال پیام توسط ادمین الزامی است"}), 400
        # Optional: Verify engineer exists
        engineers = load_engineers()
        if not any(e['name'] == engineer_name for e in engineers):
             return jsonify({"status": "error", "message": "مهندس مشخص شده یافت نشد"}), 404
    else:
        return jsonify({"status": "error", "message": "نقش کاربر نامعتبر است"}), 403

    try:
        new_msg = add_message(engineer_name, sender_role, content)
        # Optionally notify the other party (e.g., using a simple notification system or just rely on polling/refresh)
        return jsonify({"status": "success", "message": "پیام ارسال شد", "sent_message": new_msg}), 201
    except Exception as e:
        print(f"Error sending message: {e}")
        return jsonify({"status": "error", "message": "خطا در ارسال پیام"}), 500

@app.route('/api/messages/<engineer_name>', methods=['GET'])
@login_required
def get_messages_api(engineer_name):
    user = session['user']
    # Ensure only the relevant engineer or an admin can access messages
    if user['role'] == 'engineer' and user.get('engineer_name') != engineer_name:
        return jsonify({"status": "error", "message": "دسترسی غیرمجاز"}), 403
    if user['role'] not in ['admin', 'engineer']:
         return jsonify({"status": "error", "message": "نقش نامعتبر"}), 403

    messages = get_messages_for_engineer(engineer_name)
    return jsonify({"status": "success", "messages": messages})

@app.route('/api/messages/engineer/markread/<engineer_name>', methods=['POST'])
@engineer_required
def mark_engineer_read_api(engineer_name):
    user = session['user']
    if user.get('engineer_name') != engineer_name:
        return jsonify({"status": "error", "message": "دسترسی غیرمجاز"}), 403

    updated = mark_engineer_messages_as_read(engineer_name)
    return jsonify({"status": "success", "marked_read": updated})

@app.route('/api/messages/admin/markread/<engineer_name>', methods=['POST'])
@admin_required
def mark_admin_read_api(engineer_name):
    # Optional: Check if engineer_name is valid
    updated = mark_admin_messages_as_read(engineer_name)
    return jsonify({"status": "success", "marked_read": updated})

# --- END Messaging API Routes ---

# --- New API Route for Overall Excel ---
@app.route('/api/generate_overall_excel', methods=['POST'])
@login_required
def generate_overall_excel():
    data = request.json
    year = data.get('year')
    month = data.get('month')

    if not year or not month:
        return jsonify({"status": "error", "error": "Year and month are required."}), 400

    schedules = load_schedules()
    period_key = f"{year}-{month}"
    schedule_data_for_period = schedules.get(period_key)

    if not schedule_data_for_period:
        return jsonify({"status": "error", "error": f"No schedule data found for {year}-{month}."}), 404

    # Use Jalali year/month in filename
    filename = f"Overall_Shifts_{year}_{month}.xlsx"
    file_path = os.path.join(DATA_DIR, filename)

    try:
        create_overall_excel_schedule(file_path, str(year), str(month), schedule_data_for_period)
        return jsonify({"status": "success", "file": filename})
    except Exception as e:
        logging.error(f"Error generating overall Excel for {year}-{month}: {e}", exc_info=True)
        return jsonify({"status": "error", "error": f"Failed to generate overall Excel: {str(e)}"}), 500

# --- END New API Route ---

# --- Engineer Panel Lock Routes ---
@app.route('/admin/toggle_engineer_lock', methods=['POST'])
@admin_required
def toggle_engineer_lock():
    global ENGINEER_PANEL_LOCKED
    ENGINEER_PANEL_LOCKED = not ENGINEER_PANEL_LOCKED
    status_message = "قفل شد" if ENGINEER_PANEL_LOCKED else "باز شد"
    flash(f"دسترسی مهندسان برای ثبت محدودیت {status_message}.", "success" if not ENGINEER_PANEL_LOCKED else "warning")
    return jsonify({"status": "success", "locked": ENGINEER_PANEL_LOCKED, "message": f"Engineer panel is now {'locked' if ENGINEER_PANEL_LOCKED else 'unlocked'}."})

@app.route('/api/engineer_lock_status', methods=['GET'])
@login_required # Should be accessible by engineers too
def get_engineer_lock_status():
    global ENGINEER_PANEL_LOCKED, LOCK_MESSAGE
    return jsonify({"locked": ENGINEER_PANEL_LOCKED, "lock_message": LOCK_MESSAGE if ENGINEER_PANEL_LOCKED else ""})
# --- End Engineer Panel Lock Routes ---

if __name__ == '__main__':
    app.run(debug=True, port=8000)